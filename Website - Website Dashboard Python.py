import os
import pandas as pd
import pyodbc
from html import unescape
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
# Table functionality was removed as it caused Excel recovery errors
from datetime import datetime
from rapidfuzz import fuzz

# Paths to CSV files
BASE_DIR = r'C:\Users\dmccrea\Documents\Python Scripts\New folder'
PRICING_CSV = os.path.join(BASE_DIR, 'AU_product_offer_price_en_AU.csv')
PUBLISHED_OFFERS_CSV = os.path.join(BASE_DIR, 'publishedOffers.csv')
PRODUCTS_CSV = os.path.join(BASE_DIR, 'AU_products_en_AU.csv')
IMAGES_CSV = os.path.join(BASE_DIR, 'AU_product_image_en_AU.csv')
SAP_LISTINGS_XLSX = os.path.join(BASE_DIR, 'SAP Listings.xlsx')

# SQL Server connection parameters
SERVER = '5909z0ndbsrvt02'
DATABASE = 'BIRD_IDS_DDS'
CONN_STR = (
    'DRIVER={ODBC Driver 17 for SQL Server};'
    f'SERVER={SERVER};'
    f'DATABASE={DATABASE};'
    'Trusted_Connection=yes;'
    'Encrypt=yes;'
    'TrustServerCertificate=yes;'
)

EXPECTED_REGIONS = ['BRE', 'DAN', 'DER', 'JKT', 'MIN', 'PRE', 'RGY', 'STP']

# Template for creating a Buying Smart Search (BSS) link. Only the first
# six digits of the sellable ID are required for the URL.
BSS_LINK_TEMPLATE = (
    "https://apacregional-iit-au-prd.launchpad.cfapps.ap20.hana.ondemand.com/"
    "site?siteId=b8480577-b94c-40bd-8ab2-4131ea48f108"
    "#BuyingSmartSearch-manage?sap-ui-app-id-hint=saas_approuter_com.sap.aldi.BuyingSmartSearch"
    "&/detail/{id}/TwoColumnsMidExpanded"
)


def build_bss_link(sellable_id):
    """Return BSS product link for the given sellable ID."""
    try:
        first_six = str(int(sellable_id))[:6]
    except (ValueError, TypeError):
        return None
    return BSS_LINK_TEMPLATE.format(id=first_six)


def build_website_link(sellable_id):
    """Return website product link for the given sellable ID."""
    try:
        return f"https://www.aldi.com.au/product/{int(sellable_id):018d}"
    except (ValueError, TypeError):
        return None


def check_file(path, description):
    """Print whether the given file path exists."""
    if os.path.exists(path):
        print(f"{description} found: {path}")
        return True
    else:
        print(f"{description} NOT FOUND: {path}")
        return False


def load_region_lookup(conn):
    query = """
    SELECT [AHEAD_Plant_ID] AS StoreID, [Legacy_Region_Name_Short] AS Region
    FROM dds.INT_OBJ_MD_Store
    """
    df = pd.read_sql(query, conn)
    return dict(zip(df['StoreID'].astype(str), df['Region']))


def load_store_data(conn):
    """Return dataframe of store metadata including names and regions."""
    query = """
    SELECT [AHEAD_Plant_ID] AS StoreID,
           [AHEAD_Store_Name] AS StoreName,
           [Legacy_Region_Name_Short] AS Region
    FROM dds.INT_OBJ_MD_Store
    """
    df = pd.read_sql(query, conn)
    df['StoreID'] = df['StoreID'].astype(str)
    return df


def load_pricing_data():
    df = pd.read_csv(
        PRICING_CSV,
        usecols=['concrete_sku', 'merchant_reference', 'value_gross', 'is_active'],
        dtype={'merchant_reference': 'string', 'is_active': 'string'}
    )
    df['concrete_sku'] = pd.to_numeric(df['concrete_sku'], errors='coerce').astype('Int64')
    df['value_gross'] = pd.to_numeric(df['value_gross'], errors='coerce').astype('Int64')
    df = df[df['is_active'] == '1'].copy()
    df = df.rename(columns={
        'concrete_sku': 'SellableID',
        'merchant_reference': 'StoreID',
        'value_gross': 'RetailCents'
    })
    return df[['SellableID', 'StoreID', 'RetailCents']]


def load_published_offers():
    df = pd.read_csv(PUBLISHED_OFFERS_CSV, usecols=['offer reference'], dtype=str)
    parts = df['offer reference'].str.split('_', n=1, expand=True)
    df['SellableID'] = pd.to_numeric(parts[0].str.lstrip('0'), errors='coerce').astype('Int64')
    df['StoreID'] = parts[1].str.lstrip('0')
    df = df.dropna(subset=['SellableID', 'StoreID'])
    return df[['SellableID', 'StoreID']]


def compute_product_pricing(pricing_df, region_map):
    df = pricing_df.copy()
    df['Retail'] = df['RetailCents'] / 100.0
    df['Region'] = df['StoreID'].astype(str).map(region_map)
    df = df.dropna(subset=['Region'])
    df = df.drop_duplicates(subset=['SellableID', 'Retail', 'Region'])

    grouped = (
        df.sort_values('Region')
          .groupby(['SellableID', 'Retail'])['Region']
          .apply(lambda r: ', '.join(r))
          .reset_index(name='RegionList')
    )

    def flag_all(regions):
        regions_sorted = sorted(regions.split(', '))
        return 'ALL' if regions_sorted == EXPECTED_REGIONS else regions

    grouped['Regions'] = grouped['RegionList'].apply(flag_all)
    return grouped[['SellableID', 'Retail', 'Regions']]


def compute_product_location(offers_df, region_map):
    df = offers_df[['SellableID', 'StoreID']].drop_duplicates()
    df['Region'] = df['StoreID'].astype(str).map(region_map)
    df = df.dropna(subset=['Region'])

    grouped = df.groupby('SellableID').agg(
        Regions=('Region', lambda x: ', '.join(sorted(set(x)))),
        StoreCount=('StoreID', 'nunique')
    ).reset_index()

    def flag_all(regions):
        return 'ALL' if sorted(regions.split(', ')) == EXPECTED_REGIONS else regions

    grouped['Regions'] = grouped['Regions'].apply(flag_all)
    return grouped


def compute_intra_region_price_variation(pricing_df, store_df, valid_stores):
    """Identify price differences within each region for a product.

    Only stores present in ``valid_stores`` are considered. The output samples
    include the store name together with the store id and region similar to the
    listing discrepancy sheet.
    """

    df = (
        pricing_df[pricing_df['StoreID'].astype(str).isin(valid_stores)]
        .merge(
            store_df[['StoreID', 'StoreName', 'Region']],
            on='StoreID',
            how='left'
        )
    )
    df['Retail'] = df['RetailCents'] / 100.0
    df = df.dropna(subset=['Region'])

    records = []
    for (sid, region), grp in df.groupby(['SellableID', 'Region']):
        if grp['Retail'].nunique() > 1:
            samples = []
            for price, sub in grp.groupby('Retail'):
                info = [
                    f"{row.StoreName} ({row.StoreID}, {row.Region})"
                    for row in sub[['StoreName', 'StoreID', 'Region']].itertuples(index=False)
                ]
                samples.append(f"{price:.2f}: {', '.join(info[:2])}")
            records.append({
                'Sellable ID': sid,
                'Region': region,
                'Store Price Sample': '; '.join(samples)
            })
    return pd.DataFrame(records)


def load_product_catalog():
    df = pd.read_csv(
        PRODUCTS_CSV,
        usecols=[
            'concrete_sku', 'is_active', 'name', 'description', 'brand_name',
            'price_unit', 'comparison_price_unit', 'net_content', 'content_unit',
            'product_class', 'legal_disclaimer'
        ],
        dtype=str
    )

    df['concrete_sku'] = df['concrete_sku'].str.lstrip('0')
    df['Sellable ID'] = pd.to_numeric(df['concrete_sku'], errors='coerce').astype('Int64')
    df['Net Content'] = df['net_content'].fillna('').astype(str) + ' ' + df['content_unit'].fillna('').astype(str)

    df = df.rename(columns={
        'is_active': 'Online Active',
        'name': 'Product Name',
        'description': 'Product Description',
        'brand_name': 'Brand',
        'price_unit': 'Unit',
        'comparison_price_unit': 'Comparison Unit Price',
        'product_class': 'Hierarchy',
        'legal_disclaimer': 'Legal Disclaimer'
    })

    # Exclude products not active online
    df = df[df['Online Active'] == '1'].copy()

    df = df[[
        'Sellable ID', 'Online Active', 'Product Name', 'Product Description',
        'Brand', 'Unit', 'Comparison Unit Price', 'Net Content', 'Hierarchy',
        'Legal Disclaimer'
    ]]

    # Clean HTML from descriptions and decode entities
    df['Product Description'] = (
        df['Product Description']
        .astype(str)
        .str.replace(r'<[^>]+>', '', regex=True)
        .apply(unescape)
    )
    return df


def load_general_product_info(conn):
    query = """
        SELECT p.[Article] as SellableID,
               p.[Medium_Description] as Description,
               pg.short_description as BD,
               ph.Long_Description as Hierarchy,
               cg.Medium_Description as CG,
               scg.Medium_Description as SCG
        FROM [BIRD_IDS_DDS].[dds].[INT_OBJ_MD_Product] as P
        left join [BIRD_IDS_DDS].[dds].[INT_OBJ_MD_PurchasingGroup] as PG
               on pg.purchasing_group = p.purch_grp_buyer
        left join [BIRD_IDS_DDS].[dds].[INT_OBJ_MD_ProductHierarchy] as PH
               on ph.Product_Hierarchy = p.Product_Hierarchy
        left join [BIRD_IDS_DDS].[dds].[INT_OBJ_MD_SubCommodityGroup] as SCG
               on scg.Sub_Commodity_Group = p.Sub_Commodity_Group
        left join [BIRD_IDS_DDS].[dds].[INT_OBJ_MD_CommodityGroup] as CG
               on cg.Commodity_Group = p.Commodity_Group
    """
    df = pd.read_sql(query, conn)
    df['SellableID'] = pd.to_numeric(df['SellableID'], errors='coerce').astype('Int64')
    return df


def load_product_images():
    """Return dataframe of sellable IDs that have images online."""
    try:
        df = pd.read_csv(IMAGES_CSV, usecols=[0], dtype=str)
    except FileNotFoundError:
        return pd.DataFrame(columns=['SellableID'])

    df = df.rename(columns={df.columns[0]: 'concrete_sku'})
    df['concrete_sku'] = df['concrete_sku'].str.lstrip('0')
    df['SellableID'] = pd.to_numeric(df['concrete_sku'], errors='coerce').astype('Int64')
    return df[['SellableID']].dropna()


def load_sap_store_counts():
    """Return DataFrame with SAP store count and sample stores per sellable ID."""
    if not os.path.exists(SAP_LISTINGS_XLSX):
        return pd.DataFrame(columns=['SellableID', 'SAP_Count', 'StoreSample'])

    df = pd.read_excel(SAP_LISTINGS_XLSX, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    if not df.columns.empty:
        df.rename(columns={df.columns[0]: 'ProductCode'}, inplace=True)

    store_cols = [c for c in df.columns if 'store' in c.lower()]
    if not store_cols:
        return pd.DataFrame(columns=['SellableID', 'SAP_Count', 'StoreSample'])

    def parse_cell(val):
        if pd.isna(val):
            return []
        val = str(val).replace('\n', ',').replace(';', ',')
        return [s.strip().lstrip('0') for s in val.split(',') if s.strip()]

    df['StoreList'] = (
        df[store_cols]
        .apply(lambda r: [s for col in store_cols for s in parse_cell(r[col])], axis=1)
    )
    df['SAP_Count'] = df['StoreList'].apply(lambda x: len(set(x)))
    df['StoreSample'] = df['StoreList'].apply(lambda x: ', '.join(x[:10]))
    df['SellableID'] = pd.to_numeric(df['ProductCode'].astype(str).str.lstrip('0'), errors='coerce').astype('Int64')
    return df[['SellableID', 'SAP_Count', 'StoreSample', 'StoreList']]


def build_dashboard(df_catalog, df_location, df_gp, df_price, df_images):
    df = df_catalog.merge(
        df_location,
        left_on='Sellable ID',
        right_on='SellableID',
        how='left'
    ).drop(columns=['SellableID'])

    df = df.rename(columns={
        'Regions': 'Available Online by Region',
        'StoreCount': 'Available in Stores (Count)'
    })
    df['Available Online by Region'] = df['Available Online by Region'].fillna('Not Online')
    df['Available in Stores (Count)'] = df['Available in Stores (Count)'].fillna(0).astype(int)


    df = df.merge(
        df_gp,
        left_on='Sellable ID',
        right_on='SellableID',
        how='left'
    ).drop(columns=['SellableID'])

    df = df.rename(columns={
        'Description': 'SAP Description',
        'BD': 'SAP BD',
        'Hierarchy_y': 'SAP Hierarchy',
        'CG': 'SAP Commodity Group',
        'SCG': 'SAP Sub Commodity Group'
    })

    # If both Hierarchy columns exist after merge, prefer SAP Hierarchy but keep first
    if 'Hierarchy_x' in df.columns:
        df = df.drop(columns=['Hierarchy_x'])

    df = df.merge(
        df_price,
        left_on='Sellable ID',
        right_on='SellableID',
        how='left'
    ).drop(columns=['SellableID'])

    df = df.merge(
        df_images.assign(ImageStatus='Image Online'),
        left_on='Sellable ID',
        right_on='SellableID',
        how='left'
    ).drop(columns=['SellableID'])
    df = df.rename(columns={
        'Retail': 'Product Pricing.Retail',
        'Regions': 'Product Pricing.Regions'
    })

    grouped = []
    for sid, grp in df.groupby('Sellable ID'):
        first_row = grp.iloc[0]
        regions_list = ', '.join(sorted(set(grp['Product Pricing.Regions'].dropna())))
        if regions_list:
            region_label = 'ALL' if sorted(regions_list.split(', ')) == EXPECTED_REGIONS else regions_list
        else:
            region_label = ''
        price_pairs = grp[['Product Pricing.Regions', 'Product Pricing.Retail']].dropna()
        price_pairs = price_pairs.drop_duplicates().sort_values([
            'Product Pricing.Regions', 'Product Pricing.Retail'
        ])
        retail_by_region = ', '.join(
            f"{r}: {p}" for r, p in zip(
                price_pairs['Product Pricing.Regions'],
                price_pairs['Product Pricing.Retail']
            )
        )
        multiple_prices = price_pairs['Product Pricing.Retail'].nunique() > 1
        grouped.append({
            'Sellable ID': sid,
            'Product Name': first_row['Product Name'],
            'Product Description': first_row['Product Description'],
            'Brand': first_row['Brand'],
            'SAP Description': first_row.get('SAP Description'),
            'Net Content': first_row['Net Content'],
            'SAP BD': first_row.get('SAP BD'),
            'Hierarchy': first_row.get('SAP Hierarchy') or first_row.get('Hierarchy'),
            'SAP Commodity Group': first_row.get('SAP Commodity Group'),
            'SAP Sub Commodity Group': first_row.get('SAP Sub Commodity Group'),
            'Legal Disclaimer': first_row.get('Legal Disclaimer'),
            'Image Status': first_row.get('ImageStatus', 'No Image Online'),
            'Available in Stores (Count)': first_row['Available in Stores (Count)'],
            'Regions On Website': region_label,
            'Retail by Region': retail_by_region,
            'Multiple Prices': multiple_prices
        })

    out = pd.DataFrame(grouped)
    out = out.rename(columns={
        'Product Name': 'Website Product Name',
        'SAP Description': 'SAP Product Name',
        'Retail by Region': 'Retail by Region (updated weekly)'
    })
    out = out[[
        'SAP BD', 'Sellable ID', 'Website Product Name', 'SAP Product Name',
        'Regions On Website', 'Available in Stores (Count)',
        'Retail by Region (updated weekly)', 'Product Description',
        'Legal Disclaimer', 'Image Status', 'Hierarchy',
        'SAP Commodity Group', 'SAP Sub Commodity Group',
        'Brand', 'Net Content', 'Multiple Prices'
    ]]
    out['Image Status'] = out['Image Status'].fillna('No Image Online')
    return out


def main():
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(BASE_DIR, f'Website_Dashboard_Output_{timestamp}.xlsx')
    os.makedirs(BASE_DIR, exist_ok=True)
    # Check that required input files are present
    check_file(PRICING_CSV, 'Pricing CSV')
    check_file(PUBLISHED_OFFERS_CSV, 'Published Offers CSV')
    check_file(PRODUCTS_CSV, 'Products CSV')
    check_file(IMAGES_CSV, 'Images CSV')
    check_file(SAP_LISTINGS_XLSX, 'SAP listings file')
    try:
        conn = pyodbc.connect(CONN_STR)
        store_data = load_store_data(conn)
        region_map = dict(zip(store_data['StoreID'], store_data['Region']))
        store_info_map = store_data.set_index('StoreID')[['StoreName', 'Region']].to_dict('index')
        pricing_base = load_pricing_data()
        offers_base = load_published_offers()
        website_store_map = (
            offers_base.groupby('SellableID')['StoreID']
            .apply(lambda x: sorted(set(x.astype(str).str.lstrip('0'))))
            .to_dict()
        )
        pricing_data = compute_product_pricing(pricing_base, region_map)
        location_data = compute_product_location(offers_base, region_map)
        catalog = load_product_catalog()
        gp_info = load_general_product_info(conn)
        images_df = load_product_images()
        sap_counts = load_sap_store_counts()
        valid_stores = set(offers_base['StoreID'].astype(str))
        if not sap_counts.empty and 'StoreList' in sap_counts.columns:
            for lst in sap_counts['StoreList']:
                valid_stores.update(str(s) for s in lst)
        final_df = build_dashboard(catalog, location_data, gp_info, pricing_data, images_df)
        price_variation_df = compute_intra_region_price_variation(pricing_base, store_data, valid_stores)
        if not price_variation_df.empty:
            info_cols = [
                'Sellable ID',
                'Website Product Name',
                'SAP BD',
                'Hierarchy',
                'SAP Commodity Group',
                'Retail by Region (updated weekly)'
            ]
            merge_info = final_df[info_cols].drop_duplicates()
            price_variation_df = price_variation_df.merge(
                merge_info, on='Sellable ID', how='left'
            )
            other_cols = [
                'Website Product Name',
                'SAP BD',
                'Hierarchy',
                'SAP Commodity Group',
                'Retail by Region (updated weekly)'
            ]
            price_variation_df = price_variation_df.dropna(
                subset=other_cols, how='all'
            )
            price_variation_df = price_variation_df[
                [
                    'SAP BD',
                    'Sellable ID',
                    'Website Product Name',
                    'Hierarchy',
                    'Region',
                    'Retail by Region (updated weekly)',
                    'Store Price Sample',
                    'SAP Commodity Group'
                ]
            ]

        sap_store_map = dict(zip(sap_counts['SellableID'], sap_counts['StoreList']))
        final_df = final_df.merge(
            sap_counts,
            left_on='Sellable ID',
            right_on='SellableID',
            how='left'
        ).drop(columns=['SellableID'])
        final_df = final_df.rename(columns={
            'SAP_Count': 'Stores Listed in SAP',
            'StoreSample': 'SAP Store Sample',
            'StoreList': 'SAP Store List'
        })
        # Ensure blank SAP counts are treated as 0 for comparison logic
        final_df['Stores Listed in SAP'] = final_df['Stores Listed in SAP'].fillna(0).astype(int)

        # Exclude ALDI Services and gift card related articles
        exclude_mask = (
            final_df['SAP Commodity Group'].str.contains('ALDI Services', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Gift Cards', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Giftcards', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Giftcards, Vouchers, Tickets & Coupons', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Fruits', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Salads', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Vegetables', case=False, na=False)
        )
        final_df = final_df[~exclude_mask]
        if not price_variation_df.empty:
            exclude_mask_pv = (
                price_variation_df['SAP Commodity Group'].str.contains('ALDI Services', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Gift Cards', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Giftcards', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Giftcards, Vouchers, Tickets & Coupons', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Fruits', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Salads', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Vegetables', case=False, na=False)
            )
            price_variation_df = price_variation_df[~exclude_mask_pv]

        # Sort data for output
        final_df = final_df.sort_values('SAP BD')
        price_variation_df = price_variation_df.sort_values('Hierarchy')

        mismatch_counts = final_df[
            (final_df['Stores Listed in SAP'] != final_df['Available in Stores (Count)']) &
            ~((final_df['Stores Listed in SAP'] == 0) & (final_df['Available in Stores (Count)'] == 0))
        ][[
            'SAP BD', 'Sellable ID', 'Website Product Name',
            'Available in Stores (Count)', 'Stores Listed in SAP', 'SAP Store List'
        ]]

        def diff_info(primary, secondary):
            diff = [s for s in primary - secondary]
            formatted = []
            regions = set()
            for sid in diff[:5]:
                info = store_info_map.get(sid)
                if info:
                    formatted.append(
                        f"{info['StoreName']} ({sid}, {info['Region']})"
                    )
                    regions.add(info['Region'])
                else:
                    formatted.append(str(sid))
            return ', '.join(formatted), ', '.join(sorted(regions))

        cols = mismatch_counts['Sellable ID'].apply(
            lambda sid: diff_info(set(website_store_map.get(sid, [])), set(sap_store_map.get(sid, [])))
        )
        mismatch_counts[['Stores on Website Without Listing (up to 5)', 'Regions with Website Only']] = pd.DataFrame(cols.tolist(), index=mismatch_counts.index)

        cols = mismatch_counts['Sellable ID'].apply(
            lambda sid: diff_info(set(sap_store_map.get(sid, [])), set(website_store_map.get(sid, [])))
        )
        mismatch_counts[['Stores Listed without Product Available Online (up to 5)', 'Regions with SAP Only']] = pd.DataFrame(cols.tolist(), index=mismatch_counts.index)
        mismatch_counts = mismatch_counts.drop(columns=['SAP Store List'])
        mismatch_counts = mismatch_counts.sort_values('Stores Listed in SAP')
        final_df = final_df.drop(columns=['SAP Store Sample', 'SAP Store List'])

        # Determine most common store count per region combination
        mode_map = (final_df.groupby('Regions On Website')['Available in Stores (Count)']
                    .agg(lambda x: x.mode().iat[0] if not x.mode().empty else None))
        final_df['Deviation'] = final_df.apply(
            lambda r: r['Available in Stores (Count)'] != mode_map.get(r['Regions On Website']),
            axis=1
        )

        # Calculate fuzzy match score between website and SAP names
        final_df['Fuzzy Score'] = final_df.apply(
            lambda r: fuzz.ratio(
                (r['Website Product Name'] or '').lower(),
                (r['SAP Product Name'] or '').lower()
            ),
            axis=1
        )
        mismatch_df = final_df[
            final_df['Fuzzy Score'] < 30
        ][['SAP BD', 'Sellable ID', 'Website Product Name', 'SAP Product Name', 'Fuzzy Score']]
        mismatch_df = mismatch_df.sort_values('Fuzzy Score')
        mismatch_df = mismatch_df.rename(columns={'Fuzzy Score': 'Fuzzy Score %'})

        # Remove helper column from main output
        final_df = final_df.drop(columns=['Fuzzy Score'])

        # Append any SAP listings not already in the dashboard
        missing_ids = set(sap_counts['SellableID']) - set(final_df['Sellable ID'])
        if missing_ids:
            extra_info = gp_info[gp_info['SellableID'].isin(missing_ids)]
            extra_counts = sap_counts[sap_counts['SellableID'].isin(missing_ids)]
            extra = extra_counts.merge(extra_info, on='SellableID', how='left')
            extra = extra.rename(columns={
                'SellableID': 'Sellable ID',
                'Description': 'SAP Product Name',
                'BD': 'SAP BD',
                'Hierarchy': 'Hierarchy',
                'CG': 'SAP Commodity Group',
                'SCG': 'SAP Sub Commodity Group',
                'SAP_Count': 'Stores Listed in SAP'
            })
            for col in final_df.columns:
                if col not in extra.columns:
                    extra[col] = 'Not Active in Website Database'
            extra = extra[final_df.columns]
            final_df = pd.concat([final_df, extra], ignore_index=True)

        # Reapply exclusions in case new rows were added
        exclude_mask = (
            final_df['SAP Commodity Group'].str.contains('ALDI Services', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Gift Cards', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Giftcards', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Giftcards, Vouchers, Tickets & Coupons', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Fruits', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Salads', case=False, na=False) |
            final_df['SAP Commodity Group'].str.contains('Vegetables', case=False, na=False)
        )
        final_df = final_df[~exclude_mask]
        if not price_variation_df.empty:
            exclude_mask_pv = (
                price_variation_df['SAP Commodity Group'].str.contains('ALDI Services', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Gift Cards', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Giftcards', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Giftcards, Vouchers, Tickets & Coupons', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Fruits', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Salads', case=False, na=False) |
                price_variation_df['SAP Commodity Group'].str.contains('Vegetables', case=False, na=False)
            )
            price_variation_df = price_variation_df[~exclude_mask_pv]

        # Append helper column for formatting.
        # Explicitly define the desired output order so that
        # "Stores Listed in SAP" occupies column **G** in the workbook.
        column_order = [
            'SAP BD',                       # A
            'Sellable ID',                 # B
            'Website Product Name',        # C
            'SAP Product Name',            # D
            'Regions On Website',          # E
            'Available in Stores (Count)', # F
            'Stores Listed in SAP',        # G (requested position)
            'Retail by Region (updated weekly)',  # H
            'Product Description',         # I
            'Legal Disclaimer',            # J
            'Image Status',                # K
            'Hierarchy',                   # L
            'SAP Commodity Group',         # M
            'SAP Sub Commodity Group',     # N
            'Brand',                       # O
            'Net Content',                 # P
            'Multiple Prices',             # Q
            'Deviation'                    # R (hidden)
        ]
        final_df = final_df[column_order]

        final_df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        ws.title = 'Website Dashboard'

        # Auto filter only. Table creation removed to avoid Excel recovery errors
        ws.auto_filter.ref = ws.dimensions

        # Column widths and wrap text
        width_map = {
            'A': 16,  # SAP BD
            'B': 14,
            'C': 35,  # Website Product Name
            'D': 35,
            'E': 19,
            'F': 20,
            'G': 20,  # Stores Listed in SAP
            'H': 27,
            'I': 86,  # Product Description
            'J': 71,  # Legal Disclaimer
            'K': 15,
            'L': 15,
            'M': 15,
            'N': 15,
            'O': 15,
            'P': 15,
            'Q': 15,
            'R': 15,
        }
        for col, width in width_map.items():
            ws.column_dimensions[col].width = width
        for cell in ws['I']:
            cell.alignment = Alignment(wrap_text=True)

        # Grey out rows with no stores
        grey_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        rule = Rule(type='expression', dxf=DifferentialStyle(fill=grey_fill))
        avail_col = get_column_letter([c.value for c in ws[1]].index('Available in Stores (Count)') + 1)
        rule.formula = [f"${avail_col}2=0"]
        ws.conditional_formatting.add(f"A2:{get_column_letter(ws.max_column)}{ws.max_row}", rule)

        # Highlight count differences between website and SAP
        yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
        count_col = get_column_letter([c.value for c in ws[1]].index('Available in Stores (Count)') + 1)
        sap_col = get_column_letter([c.value for c in ws[1]].index('Stores Listed in SAP') + 1)
        dev_col = get_column_letter([c.value for c in ws[1]].index('Deviation') + 1)

        rule = Rule(type='expression', dxf=DifferentialStyle(fill=yellow_fill))
        rule.formula = [f"${count_col}2<>${sap_col}2"]
        ws.conditional_formatting.add(f"{count_col}2:{count_col}{ws.max_row}", rule)

        rule = Rule(type='expression', dxf=DifferentialStyle(fill=yellow_fill))
        rule.formula = [f"${sap_col}2<>${count_col}2"]
        ws.conditional_formatting.add(f"{sap_col}2:{sap_col}{ws.max_row}", rule)

        # Highlight multiple prices
        multi_col = get_column_letter([c.value for c in ws[1]].index('Multiple Prices') + 1)
        rule = Rule(type='expression', dxf=DifferentialStyle(fill=yellow_fill))
        rule.formula = [f"${multi_col}2"]
        price_col = get_column_letter([c.value for c in ws[1]].index('Retail by Region (updated weekly)') + 1)
        ws.conditional_formatting.add(f"{price_col}2:{price_col}{ws.max_row}", rule)

        # Red warnings
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        img_col = get_column_letter([c.value for c in ws[1]].index('Image Status') + 1)
        brand_col = get_column_letter([c.value for c in ws[1]].index('Brand') + 1)
        net_col = get_column_letter([c.value for c in ws[1]].index('Net Content') + 1)
        rule = Rule(type='expression', dxf=DifferentialStyle(fill=red_fill))
        rule.formula = [f"${img_col}2=\"No Image Online\""]
        ws.conditional_formatting.add(f"{img_col}2:{img_col}{ws.max_row}", rule)
        rule = Rule(type='expression', dxf=DifferentialStyle(fill=red_fill))
        rule.formula = [f"LEN(${brand_col}2)=0"]
        ws.conditional_formatting.add(f"{brand_col}2:{brand_col}{ws.max_row}", rule)
        rule = Rule(type='expression', dxf=DifferentialStyle(fill=red_fill))
        rule.formula = [f"LEN(${net_col}2)=0"]
        ws.conditional_formatting.add(f"{net_col}2:{net_col}{ws.max_row}", rule)

        # Hyperlinks for product and SAP names
        web_idx = [c.value for c in ws[1]].index('Website Product Name') + 1
        sap_idx = [c.value for c in ws[1]].index('SAP Product Name') + 1
        sid_idx = [c.value for c in ws[1]].index('Sellable ID') + 1
        for row in range(2, ws.max_row + 1):
            sid_val = ws.cell(row=row, column=sid_idx).value
            web_link = build_website_link(sid_val)
            if web_link:
                cell = ws.cell(row=row, column=web_idx)
                cell.hyperlink = web_link
                cell.font = Font(color='0000FF', underline='single')
            bss = build_bss_link(sid_val)
            if bss:
                cell = ws.cell(row=row, column=sap_idx)
                cell.hyperlink = bss
                cell.font = Font(color='0000FF', underline='single')

        # Hide helper columns
        ws.column_dimensions[dev_col].hidden = True
        ws.column_dimensions[multi_col].hidden = True

        # Create pivot summary sheet
        def not_online_count(group):
            mask = (
                (group['Available in Stores (Count)'] == 0) &
                (group['Stores Listed in SAP'] > 0)
            )
            not_active = group['Available in Stores (Count)'] == 'Not Active in Website Database'
            return group.loc[mask | not_active, 'Sellable ID'].nunique()

        # Exclude Special Buy articles from the BD summary pivot
        pivot_source = final_df[~final_df['Hierarchy'].str.contains('Special Buy', case=False, na=False)]
        pivot_df = (
            pivot_source.groupby('SAP BD')
            .apply(lambda g: pd.Series({
                'Product_Count': g['Sellable ID'].nunique(),
                'No_Image_Count': (g['Image Status'] == 'No Image Online').sum(),
                'Not in Stores Online, but listed to stores': not_online_count(g)
            }))
            .reset_index()
        )

        piv_ws = wb.create_sheet('BD Pivot')
        for r in dataframe_to_rows(pivot_df, index=False, header=True):
            piv_ws.append(r)
        piv_ws.auto_filter.ref = piv_ws.dimensions
        width_map = {'A':20, 'B':20, 'C':20, 'D':40}
        for col, width in width_map.items():
            piv_ws.column_dimensions[col].width = width
        piv_ws['F1'] = 'This Table excludes all Special Buy Products'

        # Sheet of name mismatches
        if not mismatch_df.empty:
            mis_ws = wb.create_sheet('Name Mismatch')
            for r in dataframe_to_rows(mismatch_df, index=False, header=True):
                mis_ws.append(r)
            mis_ws.auto_filter.ref = mis_ws.dimensions
            width_map = {'A':20, 'B':15, 'C':40, 'D':40, 'E':15}
            for col, width in width_map.items():
                mis_ws.column_dimensions[col].width = width
            fuzzy_col = get_column_letter(list(mismatch_df.columns).index('Fuzzy Score %') + 1)
            for cell in mis_ws[fuzzy_col][2:mis_ws.max_row+1]:
                cell.number_format = '0'

            # Hyperlink product names to the website and SAP names to BSS
            web_idx = list(mismatch_df.columns).index('Website Product Name') + 1
            sap_idx = list(mismatch_df.columns).index('SAP Product Name') + 1
            sid_idx = list(mismatch_df.columns).index('Sellable ID') + 1
            for row_idx, sid in enumerate(mismatch_df['Sellable ID'], start=2):
                web_link = f"https://www.aldi.com.au/product/{int(sid):018d}" if pd.notnull(sid) else None
                if web_link:
                    cell = mis_ws.cell(row=row_idx, column=web_idx)
                    cell.hyperlink = web_link
                    cell.font = Font(color='0000FF', underline='single')
                bss = build_bss_link(sid)
                if bss:
                    cell = mis_ws.cell(row=row_idx, column=sap_idx)
                    cell.hyperlink = bss
                    cell.font = Font(color='0000FF', underline='single')

        if not mismatch_counts.empty:
            samp_ws = wb.create_sheet('Listing Discrepancy')
            display_df = mismatch_counts.copy()
            display_df = display_df.rename(columns={
                'Sellable ID': 'Sellable ID (Smart Search link)',
                'Available in Stores (Count)': 'Stores on Website'
            })
            display_df = display_df.sort_values('Stores Listed in SAP')
            cols = list(display_df.columns)

            # Ensure "Regions with SAP Only" appears after "Regions with Website Only"
            # and before "Stores Listed without Product Available Online (up to 5)"
            if 'Regions with SAP Only' in cols and 'Regions with Website Only' in cols:
                cols.insert(
                    cols.index('Regions with Website Only') + 1,
                    cols.pop(cols.index('Regions with SAP Only'))
                )

            # Position "Stores Listed in SAP" immediately after the column
            # "Stores Listed without Product Available Online (up to 5)" when both exist.
            if (
                'Stores Listed in SAP' in cols and
                'Stores Listed without Product Available Online (up to 5)' in cols
            ):
                cols.insert(
                    cols.index('Stores Listed without Product Available Online (up to 5)') + 1,
                    cols.pop(cols.index('Stores Listed in SAP'))
                )
            elif 'Stores Listed in SAP' in cols and 'Regions with Website Only' in cols:
                cols.insert(
                    cols.index('Regions with Website Only') + 1,
                    cols.pop(cols.index('Stores Listed in SAP'))
                )

            display_df = display_df[cols]
            for r in dataframe_to_rows(display_df, index=False, header=True):
                samp_ws.append(r)

            samp_ws.auto_filter.ref = samp_ws.dimensions
            for col in range(1, samp_ws.max_column + 1):
                samp_ws.column_dimensions[get_column_letter(col)].width = 20

            width_overrides = {
                'Stores Listed without Product Available Online (up to 5)': 45,
                'Stores on Website Without Listing (up to 5)': 45,
                'Website Product Name': 40,
            }
            for col_name, width in width_overrides.items():
                if col_name in display_df.columns:
                    col_letter = get_column_letter(list(display_df.columns).index(col_name) + 1)
                    samp_ws.column_dimensions[col_letter].width = width

            for cell in samp_ws[1]:
                cell.alignment = Alignment(wrap_text=True)

            name_idx = list(display_df.columns).index('Website Product Name') + 1
            sid_idx = list(display_df.columns).index('Sellable ID (Smart Search link)') + 1
            for row_idx, sid in enumerate(mismatch_counts['Sellable ID'], start=2):
                web_link = build_website_link(sid)
                if web_link:
                    cell = samp_ws.cell(row=row_idx, column=name_idx)
                    cell.hyperlink = web_link
                    cell.font = Font(color='0000FF', underline='single')
                bss = build_bss_link(sid)
                if bss:
                    cell = samp_ws.cell(row=row_idx, column=sid_idx)
                    cell.hyperlink = bss
                    cell.font = Font(color='0000FF', underline='single')

            yellow_fill_ld = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
            orange_fill_ld = PatternFill(start_color='FFE4B5', end_color='FFE4B5', fill_type='solid')

            yellow_cols = ['Stores on Website',
                           'Stores on Website Without Listing (up to 5)',
                           'Regions with Website Only']
            orange_cols = ['Stores Listed in SAP',
                           'Stores Listed without Product Available Online (up to 5)',
                           'Regions with SAP Only']

            for col_name in yellow_cols:
                if col_name in display_df.columns:
                    col_letter = get_column_letter(list(display_df.columns).index(col_name) + 1)
                    rule = Rule(type='expression', dxf=DifferentialStyle(fill=yellow_fill_ld))
                    rule.formula = [f"LEN(${col_letter}1&\"\")>0"]
                    samp_ws.conditional_formatting.add(f"{col_letter}1:{col_letter}{samp_ws.max_row}", rule)

            for col_name in orange_cols:
                if col_name in display_df.columns:
                    col_letter = get_column_letter(list(display_df.columns).index(col_name) + 1)
                    rule = Rule(type='expression', dxf=DifferentialStyle(fill=orange_fill_ld))
                    rule.formula = [f"LEN(${col_letter}1&\"\")>0"]
                    samp_ws.conditional_formatting.add(f"{col_letter}1:{col_letter}{samp_ws.max_row}", rule)

        if not price_variation_df.empty:
            pv_ws = wb.create_sheet('Store Price Check')
            display_df = price_variation_df.drop(columns=['SAP Commodity Group'])
            display_df = display_df.rename(columns={'Sellable ID': 'Sellable ID (Smart Search link)'})
            display_df = display_df.sort_values('Hierarchy')
            for r in dataframe_to_rows(display_df, index=False, header=True):
                pv_ws.append(r)
            pv_ws.auto_filter.ref = pv_ws.dimensions
            width_map = {
                'A': 20,
                'B': 25,
                'C': 30,
                'D': 28,
                'E': 9,
                'F': 42,
                'G': 100,
            }
            for col_letter, width in width_map.items():
                pv_ws.column_dimensions[col_letter].width = width
            sample_col = get_column_letter(list(display_df.columns).index('Store Price Sample') + 1)
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            rule = Rule(type='expression', dxf=DifferentialStyle(fill=red_fill))
            rule.formula = [f"LEN(${sample_col}2)>0"]
            pv_ws.conditional_formatting.add(f"{sample_col}2:{sample_col}{pv_ws.max_row}", rule)

            # Hyperlinks for website and Smart Search
            web_idx = list(display_df.columns).index('Website Product Name') + 1
            sid_idx = list(display_df.columns).index('Sellable ID (Smart Search link)') + 1
            for row_idx, sid in enumerate(price_variation_df['Sellable ID'], start=2):
                web_link = build_website_link(sid)
                if web_link:
                    cell = pv_ws.cell(row=row_idx, column=web_idx)
                    cell.hyperlink = web_link
                    cell.font = Font(color='0000FF', underline='single')
                bss = build_bss_link(sid)
                if bss:
                    cell = pv_ws.cell(row=row_idx, column=sid_idx)
                    cell.hyperlink = bss
                    cell.font = Font(color='0000FF', underline='single')

        wb.save(output_path)
        print(f"Export successful! File saved to: {output_path}")
        os.startfile(output_path)
    except Exception as e:
        print('Error:', e)
    finally:
        if 'conn' in locals():
            conn.close()


if __name__ == '__main__':
    main()
